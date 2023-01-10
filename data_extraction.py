import textblob
import nltk
import json
import os
import re
import openpyxl
import pandas as pd
# download required nltk packages
nltk.download('stopwords')
nltk.download('opinion_lexicon')
nltk.download('cmudict')
from nltk.corpus import opinion_lexicon
from nltk.corpus import stopwords
from nltk.corpus import cmudict

# obtain stopwords
stop_words = stopwords.words('english')
# classify positive and negative words from nltk's opinion lexicon
positive_words = set(opinion_lexicon.positive())
negative_words = set(opinion_lexicon.negative())
# obtain cmu dictionary for finding syllables
cmu_dict = cmudict.dict()
column_data = []
# get id to url json
with open('ArticleExtractor\id_to_url.json') as json_file:
    id_to_url = json.load(json_file)

# define a function to compute the variables
def compute_variables(text):
    # tokenize the text
    tokens = nltk.word_tokenize(text)
    # remove stop words
    filtered_tokens = [token.lower() for token in tokens if token.lower() not in stop_words]

    # create a TextBlob object
    blob = textblob.TextBlob(text)
    # get positive and negative scores
    positive_word_scores = {word: 1 for word in positive_words}
    negative_word_scores = {word: -1 for word in negative_words}
    # compute the positive score
    positive_score = sum([positive_word_scores[token] for token in filtered_tokens if token in positive_word_scores])
    # compute the negative score
    negative_score = sum([negative_word_scores[token] for token in filtered_tokens if token in negative_word_scores])*-1
    # compute the polarity score
    polarity_score = (positive_score - negative_score) / ((positive_score + negative_score) + 0.000001)
    # compute the subjectivity score
    subjectivity_score = (positive_score + negative_score) / ((len(filtered_tokens) + 0.000001))
    # compute the average sentence length
    avg_sentence_length = sum([len(sentence) for sentence in blob.sentences]) / len(blob.sentences)
    # compute complex word count
    complex_word_count = 0
    for token in filtered_tokens:
        if token in cmu_dict:
            syllables = [len(list(y for y in x if y[-1].isdigit())) for x in cmu_dict[token]]
            if max(syllables) > 2:
                complex_word_count += 1
    # compute percentage of Complex words
    percent_complex_words = complex_word_count / len(filtered_tokens)
    # calculate the fog index
    fog_index = 0.4 * (avg_sentence_length + percent_complex_words)
    # calculate Average Number of Words Per Sentence
    avg_words_per_sentence = sum([len(sentence.words) for sentence in blob.sentences]) / len(blob.sentences)
    # compute word count
    cleaned_tokens = [token.strip("?,!.") for token in filtered_tokens]
    word_count = len(cleaned_tokens)
    # compute syllable count
    syllable_counts = []
    for token in cleaned_tokens:
        if token in cmu_dict:
            syllables_2 = [len(list(y for y in x if y[-1].isdigit())) for x in cmu_dict[token]]
            if token.endswith("es") or token.endswith("ed"):
                syllable_counts.append(max(syllables_2) - 1)
            else:
                syllable_counts.append(max(syllables_2))

    syllable_count_per_word = sum(syllable_counts)/len(cleaned_tokens)
    # compute personal pronouns
    pronoun_count = 0
    pronouns = re.compile(r'\b(I|we|my|ours|(?-i:us))\b')
    pronoun_count = len(pronouns.findall(text))
    # calculate average word length
    word_lengths = [len(token) for token in filtered_tokens]
    avg_word_length = sum(word_lengths) / len(filtered_tokens)
    # return the computed variables
    variables = [positive_score, negative_score, polarity_score, subjectivity_score, avg_sentence_length, percent_complex_words, fog_index, 
    avg_words_per_sentence, complex_word_count, word_count, syllable_count_per_word, pronoun_count, avg_word_length]
    return variables

# create workbook object
workbook = openpyxl.Workbook()
worksheet = workbook.active
# define column names
column_names = ['URL_ID', 'URL', 'POSITIVE SCORE',
'NEGATIVE SCORE',
'POLARITY SCORE',
'SUBJECTIVITY SCORE',
'AVG SENTENCE LENGTH',
'PERCENTAGE OF COMPLEX WORDS',
'FOG INDEX',
'AVG NUMBER OF WORDS PER SENTENCE',
'COMPLEX WORD COUNT',
'WORD COUNT',
'SYLLABLE PER WORD',
'PERSONAL PRONOUNS',
'AVG WORD LENGTH']
for i in range(len(column_names)):
    worksheet.cell(row=1, column=i+1).value = column_names[i]


# open the text files and compute the variables for each file
for file in os.listdir('ArticleExtractor\extracted_articles'):
    if file.endswith(".txt"):
        check_file = os.stat(os.path.join('ArticleExtractor\extracted_articles', file)).st_size
        if check_file == 0:
            continue
        with open((os.path.join('ArticleExtractor\extracted_articles', file)), "r") as f:
            text = f.read()

        url_id = file.split(".")[0]
        url = id_to_url[url_id]
        column_row = [url_id, url] + compute_variables(text)
        # save the output in the specified order
        column_data.append(column_row)

# sort values based on URL_ID column
df = pd.DataFrame(column_data)
df[[0]] = df[[0]].apply(pd.to_numeric)
df = df.sort_values(by=[0], ascending=True)
column_data = []
column_data = df.values.tolist()
print (column_data)
        # save the output in the specified order to xlsx file
for i in range(len(column_data)):
    for j in range(len(column_data[i])):
        worksheet.cell(row=i+2, column=j+1).value = column_data[i][j]

workbook.save("ArticleExtractor\Output Data Structure.xlsx")
        